#!/usr/local/bin python
#-*-coding: utf-8-*-

import os
import configparser
from openpyxl import load_workbook

d = {}

def showXlsxFiles(filePath):
    ct = 0

    for root, dirs, files in os.walk(filePath):
        global d

        if root == filePath:
            del root
            del dirs

            for file in files:
                if file.endswith(".xlsx"):
                    ct += 1

                    d[str(ct)] = file
            
            break

    if ct == 0:
        print("所选路径中未找到xlsx文件！\n")

    else:
        print("共找到%s个xlsx文件。列表如下：\n" % str(ct))

        i = 1
        while str(i) in d.keys():
            print("**【%s】 %s**"%(str(i), d[str(i)]))
            i += 1
        
        print("\n")
        
def getXlsxFile(key):
    global d

    if key not in d.keys():
        print("无效的序号，请核实！")
        return ""
    else:
        print("已选择文件“%s”。\n" % d[key])
        return d[key]

def showXlsxSheets(sheetList):
    ct = 0
    global d

    for sheet in sheetList:
        ct += 1

        d[str(ct)] = sheet

    if ct == 0:
        print("所选文件中未找到任何Sheet！\n")

    else:
        print("共找到%s个Sheet。列表如下：\n" % str(ct))

        i = 1
        while str(i) in d.keys():
            print("**【%s】 %s**"%(str(i), d[str(i)]))
            i += 1
        
        print("\n")    

def getXlsxSheet(key):
    global d

    if key not in d.keys():
        print("无效的序号，请核实！")
        return ""
    else:
        print("已选择Sheet“%s”。\n" % d[key])
        return d[key]

def getProcDate(key):
    global d

    if key not in d.keys():
        print("无效的序号，请核实！")
        return ""
    else:
        print("已选择日期“%s”。\n" % d[key])
        return d[key]

def getSubDirs(filePath):
    for root, dirs, files in os.walk(filePath):
        if root == filePath:
            del files
            return dirs
            break

def getSubFiles(filePath):
    res = []
    for root, dirs, files in os.walk(filePath):
        del root
        del dirs

        for file in files:
            res.append(file)
        
    return res

def isInCellRange(cellToCheck, cellRange):
    """
    to check a cell whether in a cell range
    :param cellToCheck:
    :param cellRange:
    :return:
        True : if cell in range
        False: if cell not in range
    """
    # logging.debug("cellToCheck=[%d:%d]", cellToCheck.row, cellToCheck.col_idx)
    # logging.debug("cellRange: row=[%d:%d] col=[%d:%d]",
    #              cellRange.min_row, cellRange.max_row, cellRange.min_col, cellRange.max_col)
    if (cellToCheck.row >= cellRange.min_row) and \
        (cellToCheck.row <= cellRange.max_row) and \
        (cellToCheck.col_idx >= cellRange.min_col) and \
        (cellToCheck.col_idx <= cellRange.max_col):

        return True
    else:
        return False


def getCellRangeValue(cellRange):
    """
    get cell range value -&gt; the top left cell value
    :param cellRange:
    :return:
    """
    topLeftCell = sheet.cell(row=cellRange.min_row, column=cellRange.min_col)
    topLeftCellValue = topLeftCell.value
    return topLeftCellValue

def getRealCellValue(ws, curCell):
    """
    for openpyxl, to get real value from row and column
    expecially for merged cell, will get its (same) value from top-left cell value

    :param row:
    :param column:
    :return:
    """

    realCellValue = curCell.value

    mergedCellsRangesList = ws.merged_cells.ranges
    # logging.info("mergedCellsRangesList=%s", mergedCellsRangesList)

    # Note:
    # to efficiency , we only check cell in range or not when its value is None
    # for all merged cell value is None
    if not realCellValue:
        for eachCellRange in mergedCellsRangesList:
            if isInCellRange(curCell, eachCellRange):
                cellRangeValue = getCellRangeValue(eachCellRange)
                realCellValue =  cellRangeValue
                break

    return realCellValue

def rreplace(self, old, new, *max):
    count = len(self)
    if max and str(max[0]).isdigit():
        count = max[0]
    return new.join(self.rsplit(old, count))

conf = configparser.ConfigParser()

conf.read("config.ini", encoding="utf-8")

print("**********************************************")
print("*********全库质检报告提交情况检查工具*********")
print("**********************************************\n")
print("************************作者：戎晨飞 版本：1.0\n\n")

xlsxPath = input("请输入全库质检跟踪表（xlsx格式）所在路径：（Default: 当前路径）")

if xlsxPath == "":
    xlsxPath = "./"
else:
    if not (xlsxPath.endswith("/") or xlsxPath.endswith("\\")):
        xlsxPath = xlsxPath + "/"

print("\n")

showXlsxFiles(xlsxPath)

if "1" in d.keys():
    opt = input("请输入质检报告前面【】中的序号以选择文件：（Default: 1）")

    print("\n")

    try:
        if opt != "":
            intTmp = int(opt)
        
        if opt == "":
            opt = "1"
        
        xlsx = getXlsxFile(opt)

        d.clear()

        wb = load_workbook(xlsxPath + xlsx)

        sheetList = wb.sheetnames

        showXlsxSheets(sheetList)

        opt = input("请输入Sheet名称前面【】中的序号以选择Sheet：（Default: 1）")

        print("\n")

        try:
            if opt != "":
                intTmp = int(opt)
        
            if opt == "":
                opt = "1"

            sheetName = getXlsxSheet(opt)

            d.clear()

            sheet = wb[sheetName]

            for row in sheet.rows:
                for cell in row:
                    cell.value = getRealCellValue(sheet, cell)

            print("进展情况日期如下：\n")

            col = 1

            while col <= sheet.max_column:
                if sheet.cell(row = 1, column = col).value == "进展情况":
                    d[str(col)] = sheet.cell(row = 2, column = col).value

                    print("**【%s】 %s**"%(str(col), d[str(col)]))

                col += 1

            print("\n")

            opt = input("请输入列名前面【】中的序号以选择日期：")

            print("\n")

            try:
                if opt != "":
                    intTmp = int(opt)
        
                if opt == "":
                    opt = "1"

                procDate = getProcDate(opt)

                d.clear()

                orgList = ["江苏工程一部", "江苏工程二部", "安徽工程部", "东北工程部", "西北工程部", "内蒙古工程部"]

                for org in orgList:
                    print("**检查部门“%s” ...**\n"% org)
                    
                    r = 2

                    issList = []
                    noIssList = []

                    ctIss = 0
                    ctNoIss = 0

                    while (r+1) <= sheet.max_row:
                        r += 1
                        if sheet.cell(row = r, column = 2).value != org:
                            continue
                        else:
                            l = [sheet.cell(row = r, column = 3).value, sheet.cell(row = r, column = 4).value]
                            if (sheet.cell(row = r, column = int(opt)).value == "") or (not sheet.cell(row = r, column = int(opt)).value):
                                noIssList.append(l)
                                ctNoIss += 1
                            else:
                                issList.append(l)
                                ctIss += 1
                    
                    print("%s共%s个地区完成进展填报，%s个地区未完成进展填报。列表如下：\n"% (org, str(ctIss), str(ctNoIss)))
                    strTmp = ""
                    for i in issList:
                        strTmp = strTmp + i[0] + i[1] + "、"
                    strTmp.strip(" ")
                    rreplace(strTmp, "、", "", 1)
                    print("**完成进展填报的地区：**\n%s\n" % strTmp)

                    strTmp = ""
                    for i in noIssList:
                        strTmp = strTmp + i[0] + i[1] + "、"
                    strTmp.strip(" ")
                    rreplace(strTmp, "、", "", 1)
                    print("**未完成进展填报的地区：**\n%s\n" % strTmp)

            except Exception as E:
                #print("无效的输入！\n")
                print(E)

        except:
            print("无效的输入！\n")

    except:
        print("无效的输入！\n")

chkPath = input("请输入质检报告提交路径（日期文件夹）：")

if not (chkPath.endswith("/") or chkPath.endswith("\\")):
    chkPath = chkPath + "/"

dirList = ["01-江苏工程一部", "02-江苏工程二部", "03-安徽工程部", "04-东北工程部", "05-西北工程部", "06-内蒙古工程部"]

print("\n")
    
for dir in dirList:
    dirPath = chkPath + dir + "/"
    if not os.path.exists(dirPath):
        print("**目录“%s”不存在，请核实！**\n"% dir)
    else:
        print("**检查目录“%s” ...**\n"% dir)

        cityList = getSubDirs(dirPath)

        for city in cityList:
            cityPath = dirPath + city + "/"

            print("****检查目录“%s” ...****\n"% city)

            try:
                kw = conf.get(city, "kw")
            except:
                print("未配置该城市！\n")
                continue

            pointList = kw.split(",")

            rptList = getSubFiles(cityPath)

            cmt = city

            noCmt = city

            for point in pointList:
                sw = False

                for rpt in rptList:
                    if point in city:
                        if ((point in rpt) and (".xls" in rpt)) or (("市本级" in rpt) and (".xls" in rpt)) or (("市局" in rpt) and (".xls" in rpt)):
                            sw = True
                            break
                    else:
                        if (point in rpt) and (".xls" in rpt):
                            sw = True
                            break
                
                if sw == True:
                    if point in city:
                        cmt = cmt + "市本级" + "、"
                    else:
                        cmt = cmt + point + "、"
                
                else:
                    if point in city:
                        noCmt = noCmt + "市本级" + "、"
                    else:
                        noCmt = noCmt + point + "、"
            
            cmt = cmt.rstrip("、")
            noCmt = noCmt.rstrip("、")

            if cmt != city:
                print("%s提交质检报告的区域有：%s。"%(city, cmt))

            if noCmt != city:
                print("%s未提交质检报告的区域有：%s。"%(city, noCmt))

            if cmt == city:
                print("所有区域均未提交质检报告。")
            elif noCmt == city:
                print("所有地区均已提交质检报告。")
            
            print("\n")

os.system("pause")